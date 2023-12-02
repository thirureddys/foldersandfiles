import os
import sys
import openpyxl
import pandas as pd
import xlwt
book = xlwt.Workbook()



root='C:\Python27\phoebe-6.8.0_debug-py2.6_10_amd64_nothr.ipoe'
#print ago
l=os.listdir('C:\Python27\phoebe-6.8.0_debug-py2.6_10_amd64_nothr.ipoe')
print l
wb=openpyxl.Workbook()
df = pd.DataFrame() 
for root, dirs,files in os.walk(root):
    for d in dirs:
        print d
        l=os.listdir(os.path.join(root,d))
        sheet1 = book.add_sheet(d)
        #sheet1 = wb.create_sheet(d)
        for i,e in enumerate(l):
            sheet1.write(i,0,e)
book.save('totalfiles2.xls')
