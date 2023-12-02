import os
import sys
import openpyxl
import pandas as pd
import xlwt
from xlrd import open_workbook
from xlutils.copy import copy
import random
import re

book=openpyxl.Workbook()

#book.save('total24.xlsx')
root='C:\Python27\phoebe-6.8.0'
f=os.listdir('C:\Python27\phoebe-6.8.0')
print f
ws1=book.create_sheet('main')
for r in range(0,len(f)):
    ws1.cell(row=r+1,column=1).value=f[r]
    
book.save('totalcode1210085pyfilesandfolders.xlsx')
for root, dirs,files in os.walk(root):

    for d in dirs:
        shs = book.sheetnames
        print shs
        print d
        l=os.listdir(os.path.join(root,d))
        ex=('.py')
        l=[p for p in l if (p.endswith(ex) and os.path.isdir(p))]
        shs=[str(i) for i in shs]
        #print shs
        if d in shs:
            sheet1=book.create_sheet(d+str(random.randint(0,9)))
        else:
            sheetl = book.create_sheet(d)
        for r in range(0,len(l)):
            sheetl.cell(row=r+1,column=1).value=l[r]
##            ex=('.py')
##            if l[r].endswith(ex):
##                sheetl.cell(row=r+1,column=1).value=l[r].strip()
##                
    book.save('totalcode1210085pyfilesfolders.xlsx')
    exit


