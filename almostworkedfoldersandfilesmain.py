import os
import sys
import openpyxl
import pandas as pd
import xlwt
from xlrd import open_workbook
from xlutils.copy import copy

book=openpyxl.Workbook()

#book.save('total24.xlsx')
root='C:\Python27\phoebe-6.8.0_debug-py2.6_10_amd64_nothr.ipoe'
f=os.listdir('C:\Python27\phoebe-6.8.0_debug-py2.6_10_amd64_nothr.ipoe')
print f
ws1=book.create_sheet('main')
for r in range(0,len(f)):
    ws1.cell(row=r+1,column=1).value=f[r]
    
book.save('total26.xlsx')



############## working##############################



##shs = book.sheetnames
##print shs


##################working############################
for root, dirs,files in os.walk(root):

    for d in dirs:
        shs = book.sheetnames
        print shs
        print d
        l=os.listdir(os.path.join(root,d))    
        shs=[str(i) for i in shs]
        #print shs
        if d in shs:
            sheet1=book.create_sheet(root+d)
        else:
            sheetl = book.create_sheet(d)
        for r in range(0,len(l)):
            sheetl.cell(row=r+1,column=1).value=l[r]
##        for i,e in enumerate(l):
##            sheet1.write(i,0,e)
    book.save('total26.xlsx')
    exit


